import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime

# Load the students list from text file
def load_student_list(file_path):
    with open(file_path, 'r') as file:
        students = [line.strip() for line in file.readlines()]
    return students

# Load attendance data from CSV file
def load_attendance_data(file_path):
    df = pd.read_csv(file_path)
    df['Timestamp'] = pd.to_datetime(df['Timestamp'], format='%d/%m/%Y %H:%M:%S')  # Parsing datetime
    return df

# Define lecture date ranges (Tuesday 6:00 PM - 8:00 PM)
def is_within_lecture_time(timestamp):
    lecture_start = datetime.time(18, 0)  # 6:00 PM
    lecture_end = datetime.time(20, 0)    # 8:00 PM
    lecture_day = 1  # Tuesday is day 1 in Python's weekday (Monday=0)

    if timestamp.weekday() == lecture_day:
        return lecture_start <= timestamp.time() <= lecture_end
    return False

# Process attendance based on lectures
def process_attendance(attendance_df, students, lecture_dates):
    # Create an attendance dictionary
    attendance_record = {student: {date: 0 for date in lecture_dates} for student in students}

    for _, row in attendance_df.iterrows():
        student_roll = row['Roll']
        timestamp = row['Timestamp']
        if student_roll in attendance_record:
            for date in lecture_dates:
                if timestamp.date() == date and is_within_lecture_time(timestamp):
                    attendance_record[student_roll][date] += 1

    # Convert raw attendance counts to status
    for student, record in attendance_record.items():
        for date in record:
            if attendance_record[student][date] > 2:
                attendance_record[student][date] = 3  # Proxy case (more than 2 attendances)
            elif attendance_record[student][date] == 2:
                attendance_record[student][date] = 2  # Full attendance
            elif attendance_record[student][date] == 1:
                attendance_record[student][date] = 1  # Partial attendance
            else:
                attendance_record[student][date] = 0  # Absent

    return attendance_record

# Write attendance to Excel (with required columns)
def write_to_excel(attendance_record, lecture_dates, output_file):
    data = []
    for student, record in attendance_record.items():
        total_attendance = sum([v for v in record.values()])
        total_lectures = len(lecture_dates) * 2  # Each lecture is 2 hours
        proxy = max(0, total_attendance - total_lectures)  # Calculate proxy

        # Build a row for each student
        row = [student] + [record[date] for date in lecture_dates] + [total_lectures, total_attendance, total_lectures, proxy]
        data.append(row)

    # Create DataFrame with the specified columns
    columns = ['Roll'] + [date.strftime('%d-%m-%Y') for date in lecture_dates] + ['Total count of dates', 'Total Attendance Marked', 'Total Attendance allowed', 'Proxy']
    df = pd.DataFrame(data, columns=columns)

    # Write to Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Attendance', index=False)

    # Apply conditional formatting
    apply_conditional_formatting(output_file, 'Attendance')

# Apply color coding for attendance
def apply_conditional_formatting(file_path, sheet_name):
    # Load the workbook and select the sheet
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=len(ws[1]) - 4):  # Skip last 4 columns
        for cell in row:
            if cell.value == 0:
                cell.fill = red_fill
            elif cell.value == 1:
                cell.fill = yellow_fill
            elif cell.value == 2:
                cell.fill = green_fill
            elif cell.value > 2:
                cell.fill = red_fill  # Highlight for proxy

    wb.save(file_path)

# Main function to run the processing
def main():
    # Use the file names as they appear after upload
    students_file = 'stud_list.txt'
    attendance_file = 'input_attendance.csv'
    output_file = 'outputb.xlsx'

    # Define the lecture dates for the course
    lecture_dates = [
        datetime.date(2024, 8, 6), datetime.date(2024, 8, 13), datetime.date(2024, 8, 20),
        datetime.date(2024, 8, 27), datetime.date(2024, 9, 3),
        datetime.date(2024, 9, 17), datetime.date(2024, 10, 1)
    ]

    # Load student list and attendance data
    students = load_student_list(students_file)
    attendance_data = load_attendance_data(attendance_file)

    # Process the attendance
    attendance_record = process_attendance(attendance_data, students, lecture_dates)

    # Generate the Excel report
    write_to_excel(attendance_record, lecture_dates, output_file)

if __name__== "__main__":
    main()
